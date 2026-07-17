import os, csv, statistics, urllib.request
import openpyxl

DD=r"E:\DATA\AlloTf\data"; RD=r"E:\DATA\AlloTf\results"
os.makedirs(os.path.join(DD,"tables"),exist_ok=True); TB=os.path.join(DD,"tables")
XLSX=os.path.join(DD,"MOESM11.xlsx")

# 11 design positions -> TtgR residue numbers (offset +67, verified via L92/L113)
POS_IDX=[0,3,7,11,22,25,26,29,43,46,47]
RESID=[p+67 for p in POS_IDX]   # [67,70,74,78,89,92,93,96,110,113,114]
print("design residues:",RESID)

LIGS=["Ndes","4Hy","Nar","EllA","Phlo","Tam","Nal","End","Quin"]
LIGINFO={
 "Ndes":("N-desmethyltamoxifen","CCC(=C(C1=CC=CC=C1)C2=CC=C(C=C2)OCCNC)C3=CC=CC=C3","non_native","nonhydroxylated_tamoxifen","passed"),
 "4Hy":("4-hydroxytamoxifen","CCC(=C(C1=CC=C(C=C1)O)C2=CC=C(C=C2)OCCN(C)C)C3=CC=CC=C3","non_native","hydroxylated_tamoxifen","passed"),
 "Nar":("naringenin","C1C(OC2=CC(=CC(=C2C1=O)O)O)C3=CC=C(C=C3)O","native","native_flavonoid","passed"),
 "EllA":("ellagic acid","C1=C2C3=C(C(=C1O)O)OC(=O)C4=CC(=C(C(=C43)OC2=O)O)O","non_native","ellagic_acid","failed_secondary_validation"),
 "Phlo":("phloretin","C1=CC(=CC=C1CCC(=O)C2=C(C=C(C=C2O)O)O)O","native","native_flavonoid","passed"),
 "Tam":("tamoxifen","CCC(=C(C1=CC=CC=C1)C2=CC=C(C=C2)OCCN(C)C)C3=CC=CC=C3","non_native","nonhydroxylated_tamoxifen","passed"),
 "Nal":("naltrexone","C1CC1CN2CCC34C5C(=O)CCC3(C2CC6=C4C(=C(C=C6)O)O5)O","non_native","naltrexone","passed"),
 "End":("endoxifen","CCC(=C(C1=CC=C(C=C1)O)C2=CC=C(C=C2)OCCNC)C3=CC=CC=C3","non_native","hydroxylated_tamoxifen","passed"),
 "Quin":("quinine","COC1=CC2=C(C=CN=C2C=C1)C(C3CC4CCN3CC4C=C)O","non_native","quinine","passed")}

# WT 11-mer from UniProt Q9AIU0
def wt_11mer():
    try:
        seq="".join(urllib.request.urlopen("https://rest.uniprot.org/uniprotkb/Q9AIU0.fasta",timeout=30)
                    .read().decode().split("\n")[1:])
        wt="".join(seq[r-1] for r in RESID)   # 1-indexed
        return wt, seq
    except Exception as e:
        return None, None
wt,seq=wt_11mer()
print("WT 11-mer at design residues:",wt, "| L92(idx5)=%s L113(idx9)=%s"%(wt[5] if wt else "?", wt[9] if wt else "?"))

# ---- ligands.csv ----
with open(os.path.join(TB,"ligands.csv"),"w",newline="") as f:
    w=csv.writer(f); w.writerow(["ligand_id","ligand_name","smiles","native_or_non_native","chemotype_group","secondary_validation_status"])
    for l in LIGS:
        nm,sm,nat,ct,vs=LIGINFO[l]; w.writerow([l,nm,sm,nat,ct,vs])

# ---- read Fig.2C (QC 16,191 variants, 9 ligands) ----
wb=openpyxl.load_workbook(XLSX,read_only=True,data_only=True)
ws=wb["Fig.2C"]; it=ws.iter_rows(values_only=True); hdr=list(next(it))
idx={h:i for i,h in enumerate(hdr)}
ivar=0; imut=idx.get("mut_seqs_wt")
import math
vrows=[]; rlong=[]
for r in it:
    mut=r[imut]
    if not mut or len(str(mut))!=11: continue
    vid=r[ivar]; mut=str(mut)
    mc = sum(1 for a,b in zip(mut,wt) if a!=b) if wt else ""
    vrows.append([vid,mut,mc]+list(mut))
    for l in LIGS:
        try: F=float(r[idx[l]])
        except (TypeError,ValueError,KeyError): continue
        rlong.append([vid,l,"%.5f"%F,"%.5f"%math.log2(max(F,0.05))])
wb.close()

with open(os.path.join(TB,"variants.csv"),"w",newline="") as f:
    w=csv.writer(f); w.writerow(["variant_id","mutation_string","mutation_count"]+["aa_%d"%r for r in RESID])
    w.writerows(vrows)
with open(os.path.join(TB,"responses_long.csv"),"w",newline="") as f:
    w=csv.writer(f); w.writerow(["variant_id","ligand_id","F_score","log2_F_score"]); w.writerows(rlong)
print("variants:",len(vrows)," response rows:",len(rlong))
if wt:
    mcs=[v[2] for v in vrows]; from collections import Counter
    print("mutation_count dist:",dict(sorted(Counter(mcs).items())))

# ---- variant_labels.csv from Supp.Fig.6 basal (EtOH vehicle) ----
wb=openpyxl.load_workbook(XLSX,read_only=True,data_only=True)
ws=wb["Supp. Fig.6"]; rows=list(ws.iter_rows(values_only=True)); wb.close()
basal={}
for r in rows[1:]:
    v=r[1]
    try: basal[v]=statistics.mean([float(r[2]),float(r[3]),float(r[4])])
    except (TypeError,ValueError): pass
WTb=basal.get("WT")
thr=4*WTb if WTb else statistics.quantiles(list(basal.values()),n=100)[84]
comp=sum(1 for b in basal.values() if b<thr)
print("WT basal=%.3f  competent threshold(4xWT)=%.3f  %%competent=%.1f%%"%(WTb or -1,thr,100*comp/len(basal)))
with open(os.path.join(TB,"variant_labels.csv"),"w",newline="") as f:
    w=csv.writer(f); w.writerow(["variant_id","basal_expression","repression_competent","repression_label_source","repression_confidence"])
    for v in [row[0] for row in vrows]:
        b=basal.get(v)
        if b is None:
            w.writerow([v,"NA","NA","unavailable","NA"])
        else:
            rc=1 if b<thr else 0
            conf="high" if abs(b-thr)>0.5*thr else "low"
            w.writerow([v,"%.4f"%b,rc,"Supp.Fig.6_EtOH_vehicle_basal_WTanchored",conf])
print("\nwrote 4 tables to data/tables/ : ligands, variants, responses_long, variant_labels")
